import openpyxl
import pytest as pytest


class TestExcelSheetAdd:
    def test_excelsheet(self):
        workbook = openpyxl.load_workbook("C:/Users/shashi/Desktop/ExcelSheet.xlsx")
        sheet = workbook['Login']
        i = 20
        for r in range(2, 7):
            for c in range(3, 3):
                sheet.cell(row=r, column=c).value = i
                i = i + 1
        workbook.save("C:/Users/shashi/Desktop/ExcelSheet.xlsx")
        for r in range(6,11):
            for c in range(1,4):
                sheet.cell(row=r, column=1).value = "abc" + str(i) + "@gmail.com"
                sheet.cell(row=r, column=2).value = "abc" + str(i)
                sheet.cell(row=r, column=3).value = i
                i = i + 1
        workbook.save("C:/Users/shashi/Desktop/ExcelSheet.xlsx")
