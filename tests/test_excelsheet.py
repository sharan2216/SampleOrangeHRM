import time
import openpyxl
import pytest as pytest
import selenium.webdriver.common.driver_finder
from allure_commons.types import AttachmentType
from selenium import webdriver
from selenium.webdriver.common.by import By
import tests


class TestExcelSheet:
    def test_excelsheet(self):
        workbook = openpyxl.load_workbook("C:/Users/shashi/Desktop/ExcelSheet.xlsx")
        sheet = workbook['Login']
        row_count = sheet.max_row
        col_count = sheet.max_column
        print(row_count)
        print(col_count)
        for i in range(1,row_count+1):
            for j in range(1,col_count+1):
                print(sheet.cell(i,j).value,end=" ")
            print()


