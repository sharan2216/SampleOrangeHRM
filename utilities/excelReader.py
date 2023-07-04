import pytest
import openpyxl

def get_data_from_ExcelFile(path,sheetname):
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    r = sheet.max_row
    c = sheet.max_column
    for r in range(2,r+1):
        row_list = []
        for c in range(1,c+1):
            row_list.append(sheet.cell(r, c).value)
        final_list.append(row_list)

    return final_list